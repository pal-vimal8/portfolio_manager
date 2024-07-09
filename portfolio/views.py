
# portfolio/views.py

from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from openpyxl.styles import PatternFill
import pandas as pd
from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re



def home(request):
    return render(request, 'portfolio/home.html')


def normalize_name(name):
    # Ensure name is a string, remove leading and trailing spaces, convert to lowercase, and remove extra spaces between words
    if pd.isna(name):
        return ''
    return re.sub(r'\s+', ' ', str(name).strip().lower())

def portfolio_creation(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)

        # Load the uploaded Excel file
        df = pd.read_excel(file_path)

        # Ensure only the first two columns are processed
        df = df.iloc[:, :2]

        # Normalize and clean data for comparison
        df.columns = df.columns.str.strip().str.lower()
        df = df.applymap(normalize_name)

        # Create a new workbook to write the results
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Data"

        # Define the fill colors
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Write the header
        ws.append([df.columns[0], df.columns[1]])

        # Write the data and apply the color formatting
        for index, row in df.iterrows():
            ws.append([row[0], row[1]])
            cell_a = ws[f'A{index + 2}']
            cell_b = ws[f'B{index + 2}']

            if row[0] == row[1]:
                cell_a.fill = green_fill
                cell_b.fill = green_fill
            else:
                cell_a.fill = yellow_fill
                cell_b.fill = yellow_fill

        # Identify unique names in each column
        unique_in_col1 = set(df.iloc[:, 0]) - set(df.iloc[:, 1])
        unique_in_col2 = set(df.iloc[:, 1]) - set(df.iloc[:, 0])

        # Apply yellow fill to unique names
        for index, row in df.iterrows():
            if row[0] in unique_in_col1:
                ws[f'A{index + 2}'].fill = yellow_fill
            if row[1] in unique_in_col2:
                ws[f'B{index + 2}'].fill = yellow_fill

        # Save the workbook to a HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="processed_portfolio.xlsx"'
        wb.save(response)

        return response

    return render(request, 'portfolio/portfolio_creation.html')
